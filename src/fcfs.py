import openpyxl  # For interacting with Excel files
from .sheet_operations import *
from .constants import *

def fcfs_schedule_with_requirements(file_path, tutors, rooms, subjects, groups):
    """
    Fill the timetable using FCFS scheduling, ensuring requirements are met.
    """
    # Open workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Convert subjects from tuples to dictionaries
    subject_keys = [
        'SubjectID', 'SubjectName', 'FacultyID', 'HoursRequired', 
        'RequiresBlackboard', 'RequiresComputers', 'MaxStudents', 'Type'
    ]
    room_keys = ['RoomID', 'RoomName', 'Capacity', 'HasBlackboard', 'HasComputers']
    tutor_keys = ['TutorID', 'Name', 'Surname', 'Subject']

    subjects = [dict(zip(subject_keys, subject)) for subject in subjects]
    rooms = [dict(zip(room_keys, room)) for room in rooms]
    tutors = [dict(zip(tutor_keys + [f'Schedule_{i}' for i in range(len(tutor) - len(tutor_keys))], tutor)) for tutor in tutors]

    unassigned_subjects_count = 0  # Counter for unassigned subjects

    for subject in subjects:
        subject_name = subject['SubjectName']
        faculty_id = subject['FacultyID']
        requires_blackboard = subject['RequiresBlackboard']
        requires_computers = subject['RequiresComputers']
        max_students = subject['MaxStudents']
        subject_type = subject['Type']

        # Filter suitable rooms for the subject
        suitable_rooms = [
            room for room in rooms
            if (not requires_blackboard or room['HasBlackboard']) and
               (not requires_computers or room['HasComputers']) and
               room['Capacity'] >= max_students
        ]

        if not suitable_rooms:
            print(f"No suitable rooms found for {subject_name}. Skipping.")
            unassigned_subjects_count += 1
            continue

        # Filter suitable tutors for the subject
        suitable_tutors = [tutor for tutor in tutors if tutor['Subject'] == subject_type]

        if not suitable_tutors:
            print(f"No suitable tutors found for {subject_name}. Skipping.")
            unassigned_subjects_count += 1
            continue

        scheduled = False  # To track if the subject gets scheduled

        # Try all days and times
        for day_number, day in enumerate(DAYS):
            if scheduled:
                break
            for time_number, time in enumerate(CLASS_TIMES):
                if scheduled:
                    break
                cell = f"{chr(day_number + 65)}{time_number + 2}"

                # Try every combination of room and tutor for this time slot
                for room in suitable_rooms:
                    room_sheet = workbook[f"Schedule_{room['RoomName']}"]

                    for tutor in suitable_tutors:
                        tutor_sheet = workbook[f"Schedule_{tutor['Name']}_{tutor['Surname']}"]
                        faculty_sheet = workbook[f"Schedule_{faculty_id}"]

                        # Check availability in all three schedules
                        if is_cell_occupied(tutor_sheet, cell) or \
                           is_cell_occupied(faculty_sheet, cell) or \
                           is_cell_occupied(room_sheet, cell):
                            continue

                        # Fill the schedules
                        fill_cell(tutor_sheet, cell, f"{subject_name} (Faculty {faculty_id})")
                        fill_cell(faculty_sheet, cell, f"{subject_name}")
                        fill_cell(room_sheet, cell, f"{subject_name} (Room {room['RoomName']})")

                        print(f"Scheduled {subject_name} on {day} at {time} in room {room['RoomName']} with tutor {tutor['Name']} {tutor['Surname']}.")
                        scheduled = True
                        break  # Break from tutor loop
                    if scheduled:
                        break  # Break from room loop
                if scheduled:
                    break  # Break from time loop

        if not scheduled:
            print(f"Could not schedule {subject_name}. No suitable time, room, or tutor was available.")
            unassigned_subjects_count += 1

    # Save the workbook
    workbook.save(file_path)
    print(f"All schedules updated successfully.")
    print(f"Total unassigned subjects: {unassigned_subjects_count}")


# def fcfs_schedule_with_requirements(file_path, tutors, rooms, subjects, groups):
#     """
#     Fill the timetable using FCFS scheduling, ensuring requirements are met.
#     """
#     # Open workbook
#     workbook = openpyxl.load_workbook(file_path)
    
#     no_suitable = 0

#     # Convert subjects from tuples to dictionaries
#     subject_keys = [
#         'SubjectID', 'SubjectName', 'FacultyID', 'HoursRequired', 
#         'RequiresBlackboard', 'RequiresComputers', 'MaxStudents', 'Type'
#     ]
#     room_keys = ['RoomID', 'RoomName', 'Capacity', 'HasBlackboard', 'HasComputers']
#     tutor_keys = ['TutorID', 'Name', 'Surname', 'Subject']

#     subjects = [dict(zip(subject_keys, subject)) for subject in subjects]
#     rooms = [dict(zip(room_keys, room)) for room in rooms]
#     tutors = [dict(zip(tutor_keys + [f'Schedule_{i}' for i in range(len(tutor) - len(tutor_keys))], tutor)) for tutor in tutors]

#     for subject in subjects:
#         subject_name = subject['SubjectName']
#         faculty_id = subject['FacultyID']
#         requires_blackboard = subject['RequiresBlackboard']
#         requires_computers = subject['RequiresComputers']
#         max_students = subject['MaxStudents']
#         subject_type = subject['Type']

#         # Find a suitable room
#         suitable_room = None
#         for room in rooms:
#             if (not requires_blackboard or room['HasBlackboard']) and \
#                (not requires_computers or room['HasComputers']) and \
#                room['Capacity'] >= max_students:
#                 suitable_room = room['RoomName']
#                 break

#         if not suitable_room:
#             no_suitable += 1
#             print(f"No suitable room found for {subject_name}. Skipping.")
#             continue

#         # Find a suitable tutor
#         suitable_tutor = None
#         for tutor in tutors:
#             if tutor['Subject'] == subject_type:
#                 suitable_tutor = tutor
#                 break

#         if not suitable_tutor:
#             no_suitable += 1
#             print(f"No suitable tutor found for {subject_name}. Skipping.")
#             continue

#         # Try to schedule the subject
#         for day_number, day in enumerate(DAYS):
#             for time_number, time in enumerate(CLASS_TIMES):
#                 cell = f"{chr(day_number + 65)}{time_number + 2}"
                
#                 sheet_name = "Schedule_"
#                 tutor_sheet = workbook[sheet_name + tutor['Name'] + "_" + tutor['Surname']]
#                 faculty_sheet = workbook[sheet_name + str(faculty_id)]
#                 room_sheet = workbook[sheet_name + room['RoomName']]

#                 # Check availability in all three schedules
#                 if is_cell_occupied(tutor_sheet, cell) or \
#                    is_cell_occupied(faculty_sheet, cell) or \
#                    is_cell_occupied(room_sheet, cell):
#                     continue
                
#                 # Fill the schedules
#                 fill_cell(tutor_sheet, cell, f"{subject_name} (Faculty {faculty_id})")
#                 fill_cell(faculty_sheet, cell, f"{subject_name}")
#                 fill_cell(room_sheet, cell, f"{subject_name} (Room {suitable_room})")
                
#                 print(f"Scheduled {subject_name} on {day} at {time} in room {suitable_room}.")
#                 break
#             else:
#                 # If no time was found for this day, try the next day
#                 continue
#             break  # Break out of the outer loop if scheduled
    
#     # Save the workbook
#     workbook.save(file_path)
#     print("All schedules updated successfully.")
#     print(no_suitable)



# def fcfs_schedule_with_requirements(file_path, tutors, rooms, subjects, groups):
#     """
#     Fill the timetable using FCFS scheduling, ensuring requirements are met.
    
#     Args:
#         file_path (str): Path to the Excel file.
#         sheet_name_tutor (str): Name of the tutor schedule sheet.
#         sheet_name_group (str): Name of the group schedule sheet.
#         sheet_name_room (str): Name of the room schedule sheet.
#         tasks (list of dict): Task list containing subject details and requirements.
#         rooms (list of dict): Room details with availability and resources.
#     """
#     # Open workbook
#     workbook = openpyxl.load_workbook(file_path)
    
#     # # Extract days and times from the group schedule as a reference
#     # days = [cell.value for cell in group_sheet[1][1:]]  # First row, skip first column
#     # times = [group_sheet.cell(row=i, column=1).value for i in range(2, group_sheet.max_row + 1)]
#     subject_keys = [
#         'SubjectID', 'SubjectName', 'FacultyID', 'HoursRequired', 
#         'RequiresBlackboard', 'RequiresComputers', 'MaxStudents', 'Type'
#     ]
#     room_keys = ['RoomID', 'RoomName', 'Capacity', 'HasBlackboard', 'HasComputers']
#     tutor_keys = ['TutorID', 'Name', 'Surname', 'Subject']

    
#     # Convert subjects from tuples to dictionaries
#     subjects = [dict(zip(subject_keys, subject)) for subject in subjects]
#     rooms = [dict(zip(room_keys, room)) for room in rooms]
#     tutors = [dict(zip(tutor_keys + [f'Schedule_{i}' for i in range(len(tutor) - len(tutor_keys))], tutor)) for tutor in tutors]


#     for subject in subjects:
#         subject_name = subject['SubjectName']
#         faculty_id = subject['FacultyID']
#         requires_blackboard = subject['RequiresBlackboard']
#         requires_computers = subject['RequiresComputers']
#         max_students = subject['MaxStudents']
#         subject_type = subject['Type']
        
#         # Find a suitable room
#         suitable_room = None
#         for room in rooms:
#             if (not requires_blackboard or room['HasBlackboard']) and \
#                (not requires_computers or room['HasComputers']) and \
#                room['Capacity'] >= max_students:
#                 suitable_room = room['RoomName']
#                 break

#         if not suitable_room:
#             print(f"No suitable room found for {subject_name}. Skipping.")
#             continue

#         # Find a suitable tutor
#         suitable_tutor = None
#         for tutor in tutors:
#             if tutor['Subject'] == subject_type:
#                 suitable_tutor = tutor
#                 break

#         if not suitable_tutor:
#             print(f"No suitable tutor found for {subject_name}. Skipping.")
#             continue      
        
#         # Try to schedule the subject
#         for day_number, day in enumerate(DAYS):
#             for time_number, time in enumerate(CLASS_TIMES):
#                 cell = f"{chr(day_number + 65)}{time_number + 2}"
                
#                 sheet_name = "Schedule_"
                
#                 # Check availability in all three schedules
#                 if is_cell_occupied(file_path, sheet_name + tutor['Name'] + "_" + tutor['Surname'], cell) or \
#                    is_cell_occupied(file_path, sheet_name + str(faculty_id), cell) or \
#                    is_cell_occupied(file_path, sheet_name + room['RoomName'], cell):
#                     continue
                
#                 # Fill the schedules
#                 fill_cell(file_path, sheet_name + tutor['Name'] + "_" + tutor['Surname'], cell, f"{subject_name} (Faculty {faculty_id})")
#                 fill_cell(file_path, sheet_name + str(faculty_id), cell, f"{subject_name}")
#                 fill_cell(file_path, sheet_name + room['RoomName'], cell, f"{subject_name} (Room {suitable_room})")
                
#                 print(f"Scheduled {subject_name} on {day} at {time} in room {suitable_room}.")
#                 break
#             else:
#                 # If no time was found for this day, try the next day
#                 continue
#             break  # Break out of the outer loop if scheduled
    
#     # Save the workbook
#     workbook.save(file_path)
#     print("All schedules updated successfully.")
    
#     # for subject in subjects:
#     #     subject_name = subject['SubjectName']
#     #     faculty_id = subject['FacultyID']
#     #     requires_blackboard = subject['RequiresBlackboard']
#     #     requires_computers = subject['RequiresComputers']
#     #     max_students = subject['MaxStudents']
#     #     type = subject['Type']
        
#     #     # Find a suitable room
#     #     suitable_room = None
#     #     for room in rooms:
#     #         if (not requires_blackboard or room['HasBlackboard']) and \
#     #            (not requires_computers or room['HasComputers']) and \
#     #            room['Capacity'] >= max_students:
#     #             suitable_room = room['RoomName']
#     #             break

#     #     if not suitable_room:
#     #         print(f"No suitable room found for {subject_name}. Skipping.")
#     #         continue

#     #     suitable_tutor = None
#     #     for tutor in tutors:
#     #         if tutor['Subject'] == type:
#     #             suitable_tutor = tutor
#     #             break

#     #     if not suitable_tutor:
#     #         print(f"No suitable tutor found for {subject_name}. Skipping.")
#     #         continue      
        
#     #     # Try to schedule the subject
#     #     for day_number, day  in enumerate(DAYS):
#     #         for time_number, time in enumerate(CLASS_TIMES):
#     #             cell = chr(day_number + 'A') + chr(time_number + 1)
#     #             # Get cell addresses
#     #             # tutor_cell = f"{openpyxl.utils.get_column_letter(days.index(day) + 2)}{times.index(time) + 2}"
#     #             # group_cell = tutor_cell
#     #             # room_cell = tutor_cell
#     #             sheet_name = "Schedule_"
                
#     #             # Check availability in all three schedules
#     #             if is_cell_occupied(file_path, sheet_name + tutor['Name'] + "_" + tutor['Surname'], cell) and \
#     #                is_cell_occupied(file_path, sheet_name + faculty_id, cell) and \
#     #                is_cell_occupied(file_path, sheet_name + room['RoomName'], cell):
#     #                 continue
                
#     #             # Fill the schedules
#     #             fill_cell(file_path, sheet_name + tutor['Name'] + "_" + tutor['Surname'], cell, f"{subject_name} (Faculty {faculty_id})")
#     #             fill_cell(file_path, sheet_name + faculty_id, cell, f"{subject_name}")
#     #             fill_cell(file_path, sheet_name + room['RoomName'], cell, f"{subject_name} (Room {suitable_room})")
                
#     #             print(f"Scheduled {subject_name} on {day} at {time} in room {suitable_room}.")
#     #             break
#     #         else:
#     #             # If no time was found for this day, try the next day
#     #             continue
#     #         break  # Break out of the outer loop if scheduled
    
#     # # Save the workbook
#     # workbook.save(file_path)
#     # print("All schedules updated successfully.")