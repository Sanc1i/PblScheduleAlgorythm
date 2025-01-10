import openpyxl  # For interacting with Excel files
from datetime import datetime
from .constants import *

def mark_tutor_absence(file_path, tutors):
    """
    Marks 'X' in the tutor's schedule where they are absent based on their working hours.
    """
    # Open workbook
    workbook = openpyxl.load_workbook(file_path)

    # Convert tutor data into dictionaries
    tutor_keys = ['ID', 'Name', 'Surname', 'Subject', 'FacultyID', 
                  'MondayBegin', 'TuesdayBegin', 'WednesdayBegin', 'ThursdayBegin', 'FridayBegin',
                  'MondayEnd', 'TuesdayEnd', 'WednesdayEnd', 'ThursdayEnd', 'FridayEnd']
    tutors = [dict(zip(tutor_keys, tutor)) for tutor in tutors]

    # Iterate through each tutor to mark absence
    for tutor in tutors:
        tutor_name = tutor['Name']
        tutor_surname = tutor['Surname']
        schedule_sheet_name = f"Schedule_{tutor_name}_{tutor_surname}"
        
        if schedule_sheet_name not in workbook.sheetnames:
            print(f"Sheet for {tutor_name} {tutor_surname} not found. Skipping.")
            continue

        tutor_sheet = workbook[schedule_sheet_name]
        
        for day in DAYS:
            begin_time = tutor[f"{day}Begin"]
            end_time = tutor[f"{day}End"]

            for row in range(2, tutor_sheet.max_row + 1):  # Assuming time slots start from row 2
                time_cell = tutor_sheet.cell(row=row, column=1).value  # Time slots in column A
                
                if not time_cell:
                    continue
                
                try:
                    time_obj = datetime.strptime(time_cell, "%H:%M").time()
                    begin_obj = datetime.strptime(begin_time, "%H:%M").time()
                    end_obj = datetime.strptime(end_time, "%H:%M").time()
                except ValueError:
                    continue  # Skip invalid time formats
                
                # If the time slot is outside working hours, mark with 'X'
                if time_obj < begin_obj or time_obj > end_obj:
                    day_col = DAYS.index(day) + 2  # Assuming Monday is column B (2)
                    tutor_sheet.cell(row=row, column=day_col, value='X')
            
    # Save the workbook
    workbook.save(file_path)
    print("All absences marked successfully.")