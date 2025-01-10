import os
from openpyxl import Workbook
from .constants import *


def createFile(file_name = str, namesSchedules = []): #,idx = []):
    current_folder = os.path.dirname(__file__)
    schedule_folder = os.path.join(current_folder, "..", FOLDER_NAME)
    
    if not os.path.exists(schedule_folder):
        os.makedirs(schedule_folder)
    wb = Workbook()

    # default_sheet = wb.active
    # default_sheet.title = "Schedule_1"
    # createSchedule(default_sheet)
    
    for name in (namesSchedules):
        sheet_name = f"Schedule_{name}"#_{idx}"
        new_sheet = wb.create_sheet(title=sheet_name)
        createSchedule(new_sheet)

    wb.save(get_file_path(file_name))

    print(f"Excel file has been saved at: {get_file_path(file_name)}")

def get_file_path(file_name):
    current_folder = os.path.dirname(__file__)
    schedule_folder = os.path.join(current_folder, "..", FOLDER_NAME)
    return os.path.join(schedule_folder, file_name + ".xlsx")


def createSchedule(sheet):
        for col, day in enumerate(DAYS, start=2):
            sheet.cell(row=1, column=col, value=day)

        for row, time in enumerate(CLASS_TIMES, start=2):
            sheet.cell(row=row, column=1, value=time)