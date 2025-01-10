import openpyxl
from tabulate import tabulate  # Install via `pip install tabulate`

def print_schedule(file_path, sheet_name):
    """
    Prints the schedule from a specific sheet in a tabular format.
    
    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to print.
    """
    # Open the workbook and get the sheet
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        return
    
    sheet = workbook[sheet_name]
    schedule = []

    # Read data into a list of rows
    for row in sheet.iter_rows(values_only=True):
        schedule.append(row)

    # Print the schedule in a tabular format
    print(f"\nSchedule: {sheet_name}")
    print(tabulate(schedule, headers="firstrow", tablefmt="grid"))

def print_all_schedules(file_path):
    """
    Prints schedules for all sheets in the workbook.
    
    Args:
        file_path (str): Path to the Excel file.
    """
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Print schedule for each sheet
    for sheet_name in workbook.sheetnames:
        print_schedule(file_path, sheet_name)
